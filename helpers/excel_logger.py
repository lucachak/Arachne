import os
import datetime
from openpyxl import Workbook, load_workbook

class ExcelLogger:
    def __init__(self, filename="aplicacoes_catho.xlsx"):
        self.filename = filename
        self._initialize_workbook()

    def _initialize_workbook(self):
        if not os.path.exists(self.filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "Aplicações"
            # Define headers
            ws.append(["Nome da Empresa", "Vaga", "Data de Aplicação"])
            
            # Adjust column widths for better readability
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 20
            
            wb.save(self.filename)

    def log_application(self, company_name, job_title):
        try:
            wb = load_workbook(self.filename)
            ws = wb.active
            date_str = datetime.datetime.now().strftime("%d/%m/%Y")
            ws.append([company_name, job_title, date_str])
            wb.save(self.filename)
        except Exception as e:
            print(f"Error logging to Excel: {e}")

# Global instance to be used by the scraper
excel_logger = ExcelLogger()
